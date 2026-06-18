"""
Build an Excel Sales Dashboard (.xlsx) from scratch with openpyxl.

The workbook has two sheets:
  - "Data"      : a clean transactional sales table (the single source of truth)
  - "Dashboard" : KPI cards, summary tables and native Excel charts, all driven by
                  LIVE EXCEL FORMULAS (SUM / SUMIF / SUMIFS) that reference the Data
                  sheet — change the data and every number + chart updates.

This demonstrates Excel skills the way a Business Analyst uses them: dynamic,
formula-driven reporting rather than static numbers pasted from Python.

Run:  python build_dashboard.py   ->  Sales_Dashboard.xlsx
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Sales_Dashboard.xlsx")
RNG = np.random.default_rng(2025)

FONT = "Calibri"
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D6E0F0"
WHITE = "FFFFFF"
GREY = "808080"
EUR = '#,##0\\ "€"'
PCT = "0.0%"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ----------------------------------------------------------------------
# 1. Build the source data
# ----------------------------------------------------------------------
def make_data() -> pd.DataFrame:
    n = 1_500
    regions = ["North", "South", "East", "West"]
    segments = ["Consumer", "Corporate", "Home Office"]
    cats = {"Technology": 0.16, "Furniture": 0.05, "Office Supplies": 0.13}
    base = pd.Timestamp("2024-01-01")

    cat = RNG.choice(list(cats), n, p=[0.35, 0.25, 0.40])
    price = {"Technology": 380, "Furniture": 250, "Office Supplies": 40}
    sales = np.array([RNG.gamma(2.2, price[c] / 2.2) for c in cat]).round(2)
    margin = np.array([cats[c] for c in cat])
    discount = RNG.choice([0, 0.1, 0.2], n, p=[0.6, 0.25, 0.15])
    profit = (sales * (margin - discount * 0.5) + RNG.normal(0, 6, n)).round(2)

    df = pd.DataFrame({
        "Order ID": [f"ORD-{i:05d}" for i in range(1, n + 1)],
        "Date": (base + pd.to_timedelta(RNG.integers(0, 273, n), unit="D")),
        "Region": RNG.choice(regions, n),
        "Segment": RNG.choice(segments, n, p=[0.5, 0.3, 0.2]),
        "Category": cat,
        "Quantity": RNG.integers(1, 6, n),
        "Sales": sales,
        "Profit": profit,
    }).sort_values("Date").reset_index(drop=True)
    return df


# ----------------------------------------------------------------------
# 2. Write the Data sheet
# ----------------------------------------------------------------------
def write_data(wb: Workbook, df: pd.DataFrame) -> int:
    ws = wb.create_sheet("Data")
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center")
    for _, row in df.iterrows():
        ws.append([row["Order ID"], row["Date"].to_pydatetime(), row["Region"],
                   row["Segment"], row["Category"], int(row["Quantity"]),
                   float(row["Sales"]), float(row["Profit"])])
    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(r, 2).number_format = "yyyy-mm-dd"
        ws.cell(r, 7).number_format = EUR
        ws.cell(r, 8).number_format = EUR
    widths = [12, 12, 10, 13, 16, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{last}"
    return last


# ----------------------------------------------------------------------
# 3. Build the Dashboard sheet (formula-driven)
# ----------------------------------------------------------------------
def style_title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, bold=True, size=18, color=WHITE)
    ws[cell].fill = PatternFill("solid", fgColor=NAVY)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)


def kpi_card(ws, col, label, formula, fmt):
    """Two stacked cells: label (top) + big formula-driven value (bottom)."""
    lc, vc = f"{col}3", f"{col}4"
    ws[lc] = label
    ws[lc].font = Font(name=FONT, bold=True, size=10, color=WHITE)
    ws[lc].fill = PatternFill("solid", fgColor=BLUE)
    ws[lc].alignment = Alignment(horizontal="center", vertical="center")
    ws[vc] = formula
    ws[vc].number_format = fmt
    ws[vc].font = Font(name=FONT, bold=True, size=16, color=NAVY)
    ws[vc].fill = PatternFill("solid", fgColor=LIGHT)
    ws[vc].alignment = Alignment(horizontal="center", vertical="center")
    ws[lc].border = ws[vc].border = BORDER


def table_header(ws, cells):
    for cell, text in cells:
        ws[cell] = text
        ws[cell].font = Font(name=FONT, bold=True, color=WHITE)
        ws[cell].fill = PatternFill("solid", fgColor=BLUE)
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = BORDER


def build_dashboard(wb: Workbook, df: pd.DataFrame, last: int) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([3, 16, 14, 3, 16, 14, 3, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Title ----
    ws.merge_cells("B1:I1")
    style_title(ws, "B1", "  📊  Sales Performance Dashboard  —  2024")
    ws.row_dimensions[1].height = 30

    sales_rng = f"Data!$G$2:$G${last}"
    profit_rng = f"Data!$H$2:$H${last}"
    date_rng = f"Data!$B$2:$B${last}"
    region_rng = f"Data!$C$2:$C${last}"
    cat_rng = f"Data!$E$2:$E${last}"

    # ---- KPI cards (live formulas) ----
    kpi_card(ws, "B", "Total Sales", f"=SUM({sales_rng})", EUR)
    kpi_card(ws, "C", "Total Profit", f"=SUM({profit_rng})", EUR)
    kpi_card(ws, "D", "Profit Margin", f"=C4/B4", PCT)
    kpi_card(ws, "E", "Total Orders", f"=COUNTA(Data!$A$2:$A${last})", "#,##0")
    kpi_card(ws, "F", "Avg Order Value", f"=B4/E4", EUR)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 26

    # ---- Monthly sales table (SUMIFS by month window) ----
    months = pd.period_range(df["Date"].min(), df["Date"].max(), freq="M")
    r0 = 7
    ws[f"B{r0-1}"] = "Monthly Sales"
    ws[f"B{r0-1}"].font = Font(name=FONT, bold=True, size=12, color=NAVY)
    table_header(ws, [(f"B{r0}", "Month"), (f"C{r0}", "Sales")])
    for k, p in enumerate(months):
        rr = r0 + 1 + k
        y, m = p.year, p.month
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        ws[f"B{rr}"] = p.strftime("%Y-%m")
        ws[f"C{rr}"] = (f'=SUMIFS({sales_rng},{date_rng},">="&DATE({y},{m},1),'
                        f'{date_rng},"<"&DATE({ny},{nm},1))')
        ws[f"C{rr}"].number_format = EUR
        ws[f"B{rr}"].border = ws[f"C{rr}"].border = BORDER
    month_last = r0 + len(months)

    # ---- Sales by category table (SUMIF) ----
    cats = sorted(df["Category"].unique())
    ws[f"E{r0-1}"] = "Sales by Category"
    ws[f"E{r0-1}"].font = Font(name=FONT, bold=True, size=12, color=NAVY)
    table_header(ws, [(f"E{r0}", "Category"), (f"F{r0}", "Sales")])
    for k, c in enumerate(cats):
        rr = r0 + 1 + k
        ws[f"E{rr}"] = c
        ws[f"F{rr}"] = f'=SUMIF({cat_rng},E{rr},{sales_rng})'
        ws[f"F{rr}"].number_format = EUR
        ws[f"E{rr}"].border = ws[f"F{rr}"].border = BORDER
    cat_last = r0 + len(cats)

    # ---- Profit by region table (SUMIF) ----
    regions = sorted(df["Region"].unique())
    ws[f"H{r0-1}"] = "Profit by Region"
    ws[f"H{r0-1}"].font = Font(name=FONT, bold=True, size=12, color=NAVY)
    table_header(ws, [(f"H{r0}", "Region"), (f"I{r0}", "Profit")])
    for k, rg in enumerate(regions):
        rr = r0 + 1 + k
        ws[f"H{rr}"] = rg
        ws[f"I{rr}"] = f'=SUMIF({region_rng},H{rr},{profit_rng})'
        ws[f"I{rr}"].number_format = EUR
        ws[f"H{rr}"].border = ws[f"I{rr}"].border = BORDER
    region_last = r0 + len(regions)

    # ---- Charts ----
    line = LineChart()
    line.title = "Monthly Sales Trend"
    line.style = 12
    line.height, line.width = 8, 18
    data = Reference(ws, min_col=3, min_row=r0, max_row=month_last)
    cats_ref = Reference(ws, min_col=2, min_row=r0 + 1, max_row=month_last)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats_ref)
    line.legend = None
    ws.add_chart(line, "B16")

    bar = BarChart()
    bar.title = "Sales by Category"
    bar.type, bar.style = "col", 10
    bar.height, bar.width = 8, 9
    d = Reference(ws, min_col=6, min_row=r0, max_row=cat_last)
    c = Reference(ws, min_col=5, min_row=r0 + 1, max_row=cat_last)
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(c)
    bar.legend = None
    ws.add_chart(bar, "E16")

    bar2 = BarChart()
    bar2.title = "Profit by Region"
    bar2.type, bar2.style = "col", 11
    bar2.height, bar2.width = 8, 9
    d2 = Reference(ws, min_col=9, min_row=r0, max_row=region_last)
    c2 = Reference(ws, min_col=8, min_row=r0 + 1, max_row=region_last)
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(c2)
    bar2.legend = None
    ws.add_chart(bar2, "H16")

    ws[f"B{month_last + 2}"] = ("Notes: every figure above is a live Excel formula "
                                "referencing the Data sheet — edit the data and the "
                                "dashboard recalculates automatically.")
    ws[f"B{month_last + 2}"].font = Font(name=FONT, italic=True, size=9, color=GREY)


def main() -> None:
    df = make_data()
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    last = write_data(wb, df)
    build_dashboard(wb, df, last)
    wb.calculation.fullCalcOnLoad = True  # force recalc when opened
    wb.save(OUT)

    # ---- self-check: verify the formulas would match a pandas computation ----
    print(f"Saved {OUT}")
    print("\nVerification (Excel formulas vs pandas ground truth):")
    print(f"  Total Sales : €{df['Sales'].sum():,.0f}")
    print(f"  Total Profit: €{df['Profit'].sum():,.0f}")
    print(f"  Margin      : {df['Profit'].sum()/df['Sales'].sum():.1%}")
    print(f"  Orders      : {len(df):,}")
    print(f"  AOV         : €{df['Sales'].sum()/len(df):,.0f}")
    print("  By category :", {c: round(v) for c, v in df.groupby('Category')['Sales'].sum().items()})


if __name__ == "__main__":
    main()
